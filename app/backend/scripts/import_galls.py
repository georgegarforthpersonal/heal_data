"""
Import plant gall causers from British Plant Gall Society spreadsheet.

Usage:
    ./dev-run import_galls.py              # Dry-run (preview only)
    ./dev-run import_galls.py --no-dry-run # Apply to database

Defaults to dry-run mode. Use --no-dry-run to write to database.
"""

import logging
import sys
import re
import csv
from pathlib import Path
from typing import Optional, Tuple
from dataclasses import dataclass
from datetime import datetime

from openpyxl import load_workbook

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from database.connection import get_db_cursor
from script_utils import get_arg_parser


logging.basicConfig(
    level=logging.INFO,
    format='%(levelname)s: %(message)s'
)
logger = logging.getLogger(__name__)


@dataclass
class GallCauser:
    """A gall causer species from the spreadsheet."""
    scientific_name: str
    common_name: str  # Constructed as "<type> on <host>"
    gall_type: str
    host: str
    raw_scientific: str


def extract_binomial(scientific_text: str) -> Optional[str]:
    """
    Extract just the binomial (genus + species) from scientific name.

    Examples:
        "Acalitus brevitarsus (Fockeu) (= Eriophyes brevitarsus)" -> "Acalitus brevitarsus"
        "Aceria brevipes Nalepa" -> "Aceria brevipes"
        "Meloidogyne spp." -> "Meloidogyne spp."
        "Tranzschelia anemones Persoon) Nannfeldt" -> "Tranzschelia anemones"
    """
    if not scientific_text or not isinstance(scientific_text, str):
        return None

    # Clean up the text
    text = scientific_text.strip()

    # Pattern to match genus + species (+ optional spp./sp.)
    # Genus is capitalized, species is lowercase or "spp." or "sp."
    pattern = r'^([A-Z][a-z]+)\s+([a-z]+|spp?\.|[a-z]+-[a-z]+)'
    match = re.match(pattern, text)

    if match:
        genus = match.group(1)
        species = match.group(2)
        return f"{genus} {species}"

    return None


def format_host(host_text: str) -> str:
    """
    Format host name(s) - capitalize properly (not ALL CAPS).
    Handle multiple hosts with "or".

    Examples:
        "ALNUS" -> "Alnus"
        "ANEMONE and THALICTRUM" -> "Anemone or Thalictrum"
        "CIRSIUM and SCORZONEROIDES" -> "Cirsium or Scorzoneroides"
    """
    if not host_text or not isinstance(host_text, str):
        return ""

    # Split on "and" or "," to handle multiple hosts
    hosts = re.split(r'\s+and\s+|\s*,\s*', host_text.strip())

    # Capitalize each host (first letter upper, rest lower)
    formatted_hosts = [host.strip().capitalize() for host in hosts if host.strip()]

    # Join with " or "
    return " or ".join(formatted_hosts)


def parse_excel_file(file_path: Path) -> list[GallCauser]:
    """Parse the Excel file and extract gall causer records."""
    logger.info(f"Reading Excel file: {file_path}")

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    logger.info(f"Found {ws.max_row - 1} rows of data")

    gall_causers = []
    skipped_count = 0
    seen_scientific_names = set()

    for row_idx in range(2, ws.max_row + 1):
        # Read columns
        raw_scientific = ws[f'A{row_idx}'].value
        gall_type = ws[f'E{row_idx}'].value
        host = ws[f'D{row_idx}'].value

        # Skip if any required field is missing
        if not all([raw_scientific, gall_type, host]):
            skipped_count += 1
            logger.debug(f"Row {row_idx}: Skipped - missing required field")
            continue

        # Extract scientific name
        scientific_name = extract_binomial(raw_scientific)
        if not scientific_name:
            skipped_count += 1
            logger.warning(f"Row {row_idx}: Could not extract scientific name from '{raw_scientific}'")
            continue

        # Format host
        formatted_host = format_host(host)
        if not formatted_host:
            skipped_count += 1
            logger.warning(f"Row {row_idx}: Could not format host from '{host}'")
            continue

        # Check for duplicates
        if scientific_name in seen_scientific_names:
            skipped_count += 1
            logger.debug(f"Row {row_idx}: Duplicate scientific name '{scientific_name}' - skipping")
            continue

        seen_scientific_names.add(scientific_name)

        # Construct common name: "<Gall-causer type> on <Host>"
        common_name = f"{gall_type} on {formatted_host}"

        gall_causers.append(GallCauser(
            scientific_name=scientific_name,
            common_name=common_name,
            gall_type=gall_type,
            host=formatted_host,
            raw_scientific=raw_scientific
        ))

    logger.info(f"Parsed {len(gall_causers)} gall causers")
    if skipped_count > 0:
        logger.info(f"Skipped {skipped_count} rows (missing data, invalid format, or duplicates)")

    return gall_causers


def preview_data(gall_causers: list[GallCauser], limit: int = 20):
    """Show a preview of the parsed data."""
    logger.info("\n" + "="*80)
    logger.info("DATA PREVIEW")
    logger.info("="*80)

    logger.info(f"\nShowing first {limit} records:")
    for i, gall in enumerate(gall_causers[:limit], 1):
        logger.info(f"\n{i}. {gall.common_name}")
        logger.info(f"   Scientific: {gall.scientific_name}")
        logger.info(f"   Type: {gall.gall_type}")
        logger.info(f"   Host: {gall.host}")

    if len(gall_causers) > limit:
        logger.info(f"\n... and {len(gall_causers) - limit} more")

    logger.info("\n" + "="*80)


def export_galls_csv(gall_causers: list[GallCauser], output_file: str) -> Path:
    """Export gall causers to CSV for review."""
    output_path = Path(output_file)

    with open(output_path, 'w', newline='') as f:
        writer = csv.writer(f)

        # Header
        writer.writerow([
            'Change Type',
            'Name',
            'Scientific Name',
            'Gall Type',
            'Host',
            'Raw Scientific Name'
        ])

        # All galls are new inserts
        for gall in gall_causers:
            writer.writerow([
                'INSERT',
                gall.common_name,
                gall.scientific_name,
                gall.gall_type,
                gall.host,
                gall.raw_scientific
            ])

    logger.info(f"\n✓ Gall causers exported to CSV: {output_path.absolute()}")
    return output_path


def check_existing_galls() -> int:
    """Check how many galls are already in the database."""
    with get_db_cursor() as cursor:
        cursor.execute("SELECT COUNT(*) FROM species WHERE type = 'gall'")
        count = cursor.fetchone()[0]
    return count


def import_to_database(gall_causers: list[GallCauser], dry_run: bool = True) -> bool:
    """Import gall causers to the database."""

    if dry_run:
        logger.info("\n" + "="*80)
        logger.info("DRY-RUN MODE - SUMMARY")
        logger.info("="*80)
        logger.info("No changes were made to the database.")
        logger.info("")
        logger.info(f"Would insert {len(gall_causers)} gall causers into the database.")
        logger.info("")
        logger.info("To apply these changes, run with --no-dry-run flag.")
        logger.info("="*80 + "\n")
        return True

    # Check existing
    existing_count = check_existing_galls()
    logger.info(f"\nCurrently {existing_count} galls in database")

    # Final confirmation
    logger.info("\n" + "="*80)
    logger.info("⚠️  FINAL CONFIRMATION REQUIRED")
    logger.info("="*80)
    logger.info("You are about to MODIFY THE DATABASE:")
    logger.info(f"  • Insert {len(gall_causers)} gall causer species")
    logger.info("")
    logger.info("This operation will commit changes to the database.")
    logger.info("All changes are transactional (will rollback on error).")
    logger.info("="*80)

    response = input("\nType 'yes' to confirm and proceed: ").strip().lower()
    if response != 'yes':
        logger.info("\n✗ Import cancelled by user")
        return False

    logger.info("\n" + "="*80)
    logger.info("IMPORTING TO DATABASE")
    logger.info("="*80)

    try:
        with get_db_cursor() as cursor:
            inserted_count = 0

            for gall in gall_causers:
                cursor.execute("""
                    INSERT INTO species (name, type, scientific_name)
                    VALUES (%s, %s, %s)
                """, (
                    gall.common_name,
                    'gall',
                    gall.scientific_name
                ))
                inserted_count += 1

            logger.info(f"✓ Inserted {inserted_count} gall causers")

            logger.info("\n" + "="*80)
            logger.info("✓ IMPORT SUCCESSFUL")
            logger.info("="*80)
            logger.info(f"Inserted: {inserted_count} gall causers")
            logger.info("="*80 + "\n")

            return True

    except Exception as e:
        logger.error(f"\n✗ Import failed: {e}")
        logger.error("All changes have been rolled back.")
        raise


def main(dry_run: bool = True):
    """Main script execution."""
    try:
        # Find the Excel file
        data_dir = Path(__file__).parent / "data"
        excel_files = list(data_dir.glob("British-plant-gall-causers-checklist-*.xlsx"))

        if not excel_files:
            logger.error(f"No gall causers checklist found in {data_dir}")
            return 1

        if len(excel_files) > 1:
            logger.warning(f"Multiple checklist files found, using most recent: {excel_files[-1].name}")

        excel_file = excel_files[-1]

        # Print mode banner
        if dry_run:
            logger.info("\n" + "="*80)
            logger.info("DRY-RUN MODE (Default)")
            logger.info("="*80)
            logger.info("No database changes will be made.")
            logger.info("This is a safe preview of what would happen.")
            logger.info("Use --no-dry-run to apply changes to the database.")
            logger.info("="*80 + "\n")
        else:
            logger.info("\n" + "="*80)
            logger.info("⚠️  LIVE MODE - DATABASE WRITES ENABLED")
            logger.info("="*80)
            logger.info("This will make changes to the database!")
            logger.info("Confirmation will be required.")
            logger.info("="*80 + "\n")

        # Parse the Excel file
        gall_causers = parse_excel_file(excel_file)

        if not gall_causers:
            logger.error("No valid gall causers found in file")
            return 1

        # Show preview
        preview_data(gall_causers)

        # Offer CSV export (only in non-dry-run mode)
        if not dry_run:
            logger.info("\n" + "-"*80)
            while True:
                response = input("\nExport gall causers to CSV for review? [y/n]: ").strip().lower()
                if response in ['y', 'yes']:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    csv_file = f"galls_import_preview_{timestamp}.csv"
                    csv_path = export_galls_csv(gall_causers, csv_file)
                    logger.info(f"Review the CSV at: {csv_path.absolute()}")

                    # Wait for user to review
                    input("\nPress Enter to continue after reviewing the CSV...")
                    break
                elif response in ['n', 'no']:
                    logger.info("Skipping CSV export")
                    break
                else:
                    logger.info("Invalid response. Please enter 'y' or 'n'")

        # Import to database
        success = import_to_database(gall_causers, dry_run)

        return 0 if success else 1

    except KeyboardInterrupt:
        logger.info("\n\n✗ Interrupted by user (Ctrl+C)\n")
        return 1
    except Exception as e:
        logger.error(f"\n✗ Error: {e}", exc_info=True)
        return 1


if __name__ == "__main__":
    parser = get_arg_parser(description=__doc__)
    args = parser.parse_args()

    exit_code = main(dry_run=args.dry_run)
    sys.exit(exit_code)
