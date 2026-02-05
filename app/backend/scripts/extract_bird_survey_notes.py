"""
Extract mammals and reptiles from bird survey notes and import as sightings.

Reads the bird survey Excel file and extracts species sightings from the
notes column for each survey date. Uses fuzzy matching to match species
names against mammals and reptiles in the database. Can optionally import
the matched sightings into existing Birders surveys.

Usage:
    ./staging-run extract_bird_survey_notes.py              # Dry-run (preview only)
    ./staging-run extract_bird_survey_notes.py --no-dry-run # Apply to database
    ./staging-run extract_bird_survey_notes.py -v           # Verbose output
"""

import logging
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
from rapidfuzz import fuzz, process

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from database.connection import get_db_cursor
from script_utils import get_arg_parser

logging.basicConfig(
    level=logging.INFO,
    format='%(levelname)s: %(message)s'
)
logger = logging.getLogger(__name__)


# ============================================================================
# CONFIGURATION
# ============================================================================

EXCEL_PATH = Path(__file__).parent / "data" / "Bird surveys - Heal Somerset - volunteers.xlsx"
SHEET_NAMES = ["Heal Somerset bird list 2024", "Heal Somerset bird list 2025"]

# Excel row indices (0-based)
DATE_ROW = 2
NOTES_ROW = 5

# Column indices where data starts (0=A, 1=B, 2=C)
DATA_START_COL = 2  # Column C

# Fuzzy matching threshold (0-100)
FUZZY_THRESHOLD = 70

# Survey type to match against
SURVEY_TYPE_NAME = "Birders"

# Date corrections for known typos in the spreadsheet
DATE_CORRECTIONS = {
    date(2002, 3, 15): date(2025, 3, 15),
}

# Common name aliases (normalize to DB names)
# Format: lowercase alias -> DB name (case-sensitive)
NAME_ALIASES = {
    "fox": "Red Fox",
    "hare": "Brown Hare",
    "deer": "Roe Deer",  # Default if just "deer" mentioned
    "roe": "Roe Deer",  # Handle "Roe" without "deer"
    "roe deer": "Roe Deer",
    "muntjac": "Muntjac",
    "muntjac deer": "Muntjac",
    "red deer": "Red Deer",
    "sika deer": "Sika Deer",
    "fallow deer": "Fallow Deer",
    "squirrel": "Grey Squirrel",  # Default if just "squirrel" mentioned
    "shrew": "Common Shrew",  # Default if just "shrew" mentioned
    "vole": "Field Vole",  # Default if just "vole" mentioned
    "mouse": "Wood Mouse",  # Default if just "mouse" mentioned
    "rat": "Brown Rat",  # Default if just "rat" mentioned
    "stoat": "Stoat",
    "weasel": "Weasel",
    "badger": "Badger",
    "rabbit": "Rabbit",
    "hedgehog": "Hedgehog",
    "mole": "Mole",
    "mink": "American Mink",
    "otter": "Otter",
    "slowworm": "Slow-worm",
    "slow worm": "Slow-worm",
    "slow-worm": "Slow-worm",
    "adder": "Adder",
    "grass snake": "Grass Snake",
    "common lizard": "Common Lizard",
}


# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class SpeciesSighting:
    """A species sighting parsed from notes."""
    name: str
    count: int


@dataclass
class SpeciesMatch:
    """A matched species from the database."""
    db_id: int
    db_name: str
    db_scientific_name: Optional[str]
    count: int
    original_name: str
    match_score: int


@dataclass
class NoteData:
    """Data extracted from a survey note."""
    date: date
    species_list: list[SpeciesMatch]


# ============================================================================
# PARSING FUNCTIONS
# ============================================================================

def parse_notes_for_species(notes: str) -> list[SpeciesSighting]:
    """
    Parse a notes string to extract species names and counts.

    Handles various formats:
    - "5 Roe deer (a group of 3 and a pair)"
    - "1 Fox, 2 Brown Hare, 2 Roe deer"
    - "Saw 1 Hare 4 Roe Deer"
    - "Hare 1, Roe deer 2, Shrew 1"
    - "3 Hare 3 Roe & 2 Red deer"
    - "5 Roe deer, 3Hare" (no space between number and name)

    Returns list of SpeciesSighting with name and count.
    """
    if not notes or not notes.strip():
        return []

    sightings = []

    # Remove parenthetical notes like "(a group of 3 and a pair)"
    notes = re.sub(r'\([^)]*\)', '', notes)

    # Remove common non-species phrases
    skip_phrases = [
        r'\bheard\b:?',
        r'\bsaw\b:?',
        r'\bseen\b:?',
        r'\bdead\b',
        r'\blive\b',
        r'\bonly\b',
        r'\bvery\b',
        r'\bclose\b',
        r'\bby\b',
        r'\bin\b',
        r'\bwoods\b',
        r'\bfield\b',
        r'\bfields\b',
        r'\band\b',
        r'\bwith\b',
        r'\bsome\b',
        r'\bsun\b',
        r'\bbut\b',
        r'\bnot\b',
    ]
    for phrase in skip_phrases:
        notes = re.sub(phrase, ' ', notes, flags=re.IGNORECASE)

    # Split on common delimiters: comma, period, ampersand, semicolon
    segments = re.split(r'[,;.&]+', notes)

    # Further split segments where "name count" is followed by another name
    # e.g., "mole 1  Goshawk" -> ["mole 1", "Goshawk"]
    expanded_segments = []
    for segment in segments:
        # Split where a digit is followed by whitespace and a capital letter (new species entry)
        parts = re.split(r'(\d+)\s+(?=[A-Z])', segment.strip())
        # Re-assemble: digit belongs to the preceding name
        i = 0
        while i < len(parts):
            if i + 1 < len(parts) and parts[i + 1].strip().isdigit():
                expanded_segments.append(parts[i] + parts[i + 1])
                i += 2
            else:
                expanded_segments.append(parts[i])
                i += 1

    for segment in expanded_segments:
        segment = segment.strip()
        if not segment:
            continue

        # Try to extract count and species name
        # Pattern 1: "5 Roe deer" or "5Roe deer" (number at start)
        match = re.match(r'^(\d+)\s*([A-Za-z][A-Za-z\s-]*)', segment)
        if match:
            count = int(match.group(1))
            name = match.group(2).strip()
            if name and len(name) > 1:
                sightings.append(SpeciesSighting(name=name, count=count))
            continue

        # Pattern 2: "Roe deer 5" or "Hare 1" (number at end)
        match = re.match(r'^([A-Za-z][A-Za-z\s-]*?)\s*(\d+)\s*$', segment)
        if match:
            name = match.group(1).strip()
            count = int(match.group(2))
            if name and len(name) > 1:
                sightings.append(SpeciesSighting(name=name, count=count))
            continue

        # Pattern 3: Just a species name (no count, assume 1)
        # Only if it looks like a species name (starts with capital, has letters)
        match = re.match(r'^([A-Za-z][A-Za-z\s-]+)$', segment)
        if match:
            name = match.group(1).strip()
            # Filter out very short names or common words
            if name and len(name) > 2 and name.lower() not in ['the', 'one', 'two']:
                sightings.append(SpeciesSighting(name=name, count=1))

    return sightings


def load_mammal_reptile_species(cursor) -> list[tuple]:
    """
    Load all mammal and reptile species from the database.

    Returns list of (id, name, scientific_name) tuples.
    """
    cursor.execute("""
        SELECT id, name, scientific_name
        FROM species
        WHERE type IN ('mammal', 'reptile')
        ORDER BY name
    """)
    return cursor.fetchall()


def match_species_to_db(
    sightings: list[SpeciesSighting],
    db_species: list[tuple],
    threshold: int = FUZZY_THRESHOLD
) -> tuple[list[SpeciesMatch], list[SpeciesSighting]]:
    """
    Match parsed species names against database using fuzzy matching.

    Args:
        sightings: List of SpeciesSighting from notes
        db_species: List of (id, name, scientific_name) from DB
        threshold: Minimum fuzzy match score (0-100)

    Returns:
        Tuple of (matched species list, unmatched sightings list)
    """
    if not sightings or not db_species:
        return [], sightings if sightings else []

    # Build lookup structures
    db_names = [(row[1], row[0], row[2]) for row in db_species if row[1]]  # (name, id, scientific_name)
    name_to_info = {row[1].lower(): (row[0], row[1], row[2]) for row in db_species if row[1]}

    matches = []
    unmatched = []

    for sighting in sightings:
        original_name = sighting.name
        search_name = original_name.strip()

        # Try exact match first (case-insensitive)
        if search_name.lower() in name_to_info:
            db_id, db_name, db_sci_name = name_to_info[search_name.lower()]
            matches.append(SpeciesMatch(
                db_id=db_id,
                db_name=db_name,
                db_scientific_name=db_sci_name,
                count=sighting.count,
                original_name=original_name,
                match_score=100
            ))
            continue

        # Try alias lookup
        alias_key = search_name.lower()
        if alias_key in NAME_ALIASES:
            alias_target = NAME_ALIASES[alias_key]
            if alias_target.lower() in name_to_info:
                db_id, db_name, db_sci_name = name_to_info[alias_target.lower()]
                matches.append(SpeciesMatch(
                    db_id=db_id,
                    db_name=db_name,
                    db_scientific_name=db_sci_name,
                    count=sighting.count,
                    original_name=original_name,
                    match_score=100
                ))
                continue

        # Try partial matching - if search_name contains a known species type word
        # Sort by alias length descending to prioritize more specific matches (e.g., "muntjac deer" before "deer")
        partial_match = None
        sorted_aliases = sorted(NAME_ALIASES.items(), key=lambda x: len(x[0]), reverse=True)
        for alias, target in sorted_aliases:
            if alias in search_name.lower():
                # Check if target exists in DB
                if target.lower() in name_to_info:
                    partial_match = (target, name_to_info[target.lower()])
                    break

        if partial_match:
            target_name, (db_id, db_name, db_sci_name) = partial_match
            matches.append(SpeciesMatch(
                db_id=db_id,
                db_name=db_name,
                db_scientific_name=db_sci_name,
                count=sighting.count,
                original_name=original_name,
                match_score=90  # Partial match score
            ))
            continue

        # Try fuzzy match using token_set_ratio for better partial matching
        result = process.extractOne(
            search_name,
            [x[0] for x in db_names],
            scorer=fuzz.token_set_ratio,
            score_cutoff=threshold
        )

        if result:
            matched_name, score, _ = result
            # Find the matching DB entry
            for name, db_id, db_sci_name in db_names:
                if name == matched_name:
                    matches.append(SpeciesMatch(
                        db_id=db_id,
                        db_name=name,
                        db_scientific_name=db_sci_name,
                        count=sighting.count,
                        original_name=original_name,
                        match_score=int(score)
                    ))
                    break
        else:
            unmatched.append(sighting)

    return matches, unmatched


def read_excel_and_extract_notes(verbose: bool = False) -> tuple[list[NoteData], list[tuple]]:
    """
    Read the Excel file and extract NoteData for each survey date.

    Args:
        verbose: If True, log debug information

    Returns:
        Tuple of (list of NoteData objects, list of (date, unmatched_sightings) tuples)
    """
    logger.info(f"Reading Excel file: {EXCEL_PATH}")

    if not EXCEL_PATH.exists():
        logger.error(f"Excel file not found: {EXCEL_PATH}")
        return [], []

    wb = openpyxl.load_workbook(EXCEL_PATH)
    all_note_data = []
    all_unmatched = []

    # Load DB species
    with get_db_cursor() as cursor:
        db_species = load_mammal_reptile_species(cursor)
        logger.info(f"Loaded {len(db_species)} mammal/reptile species from database")

        if db_species:
            logger.info("Sample species: " + ", ".join([s[1] for s in db_species[:5]]))

    for sheet_name in SHEET_NAMES:
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            continue

        logger.info(f"\n{'='*60}")
        logger.info(f"Processing sheet: {sheet_name}")
        logger.info('='*60)

        ws = wb[sheet_name]

        # Iterate through columns starting from DATA_START_COL
        for col_idx in range(DATA_START_COL, ws.max_column + 1):
            # Get the date from DATE_ROW
            date_cell = ws.cell(row=DATE_ROW + 1, column=col_idx + 1)  # +1 for 1-based indexing
            date_value = date_cell.value

            # Skip if no date or not a date type
            if date_value is None:
                continue

            # Convert to date object
            if isinstance(date_value, datetime):
                survey_date = date_value.date()
            elif isinstance(date_value, date):
                survey_date = date_value
            else:
                # Skip non-date values (like "Monthly Totals" columns)
                continue

            # Apply date corrections for known spreadsheet typos
            if survey_date in DATE_CORRECTIONS:
                corrected = DATE_CORRECTIONS[survey_date]
                logger.info(f"  Correcting date typo: {survey_date} -> {corrected}")
                survey_date = corrected

            # Get the notes from NOTES_ROW
            notes_cell = ws.cell(row=NOTES_ROW + 1, column=col_idx + 1)
            notes_value = notes_cell.value

            if not notes_value or not isinstance(notes_value, str):
                continue

            # Parse species from notes
            sightings = parse_notes_for_species(notes_value)

            if not sightings:
                logger.debug(f"  {survey_date}: No species found in notes")
                continue

            # Match against DB
            matches, unmatched = match_species_to_db(sightings, db_species)

            # Track unmatched for reporting
            if unmatched:
                all_unmatched.append((survey_date, notes_value, unmatched))
                if verbose:
                    for u in unmatched:
                        logger.info(f"  {survey_date}: Unmatched species: '{u.name}' (count: {u.count})")

            if matches:
                note_data = NoteData(
                    date=survey_date,
                    species_list=matches
                )
                all_note_data.append(note_data)

    return all_note_data, all_unmatched


def log_results(note_data_list: list[NoteData], unmatched_list: list[tuple]) -> None:
    """Log the extracted NoteData results."""
    logger.info("\n" + "="*80)
    logger.info("RESULTS")
    logger.info("="*80)

    if not note_data_list:
        logger.info("No data extracted.")
        return

    total_sightings = 0

    for note_data in sorted(note_data_list, key=lambda x: x.date):
        logger.info(f"\nDate: {note_data.date}")
        logger.info(f"  Species found: {len(note_data.species_list)}")
        for match in note_data.species_list:
            score_info = f" (score: {match.match_score})" if match.match_score < 100 else ""
            original_info = f" [from: '{match.original_name}']" if match.original_name.lower() != match.db_name.lower() else ""
            logger.info(f"    - {match.db_name}: {match.count}{score_info}{original_info}")
            total_sightings += match.count

    # Summary
    logger.info("\n" + "-"*80)
    logger.info("SUMMARY")
    logger.info("-"*80)
    logger.info(f"Total survey dates with mammal/reptile notes: {len(note_data_list)}")
    logger.info(f"Total species sightings: {total_sightings}")

    # Species frequency
    species_counts = {}
    for note_data in note_data_list:
        for match in note_data.species_list:
            key = match.db_name
            if key not in species_counts:
                species_counts[key] = 0
            species_counts[key] += match.count

    logger.info("\nSpecies totals across all surveys:")
    for species, count in sorted(species_counts.items(), key=lambda x: -x[1]):
        logger.info(f"  {species}: {count}")

    # Report unmatched species
    if unmatched_list:
        logger.info("\n" + "-"*80)
        logger.info("UNMATCHED SPECIES (not mammals/reptiles in DB)")
        logger.info("-"*80)
        all_unmatched_names = set()
        for survey_date, notes, unmatched in unmatched_list:
            for u in unmatched:
                all_unmatched_names.add(u.name)
        for name in sorted(all_unmatched_names):
            logger.info(f"  - {name}")


# ============================================================================
# DATABASE IMPORT FUNCTIONS
# ============================================================================

def find_birders_survey(cursor, survey_date: date) -> Optional[int]:
    """
    Find the Birders survey for a given date.

    Returns the survey ID or None if not found.
    """
    cursor.execute("""
        SELECT s.id
        FROM survey s
        JOIN survey_type st ON st.id = s.survey_type_id
        WHERE st.name = %s AND s.date = %s
    """, (SURVEY_TYPE_NAME, survey_date))
    row = cursor.fetchone()
    return row[0] if row else None


def import_sightings(note_data_list: list[NoteData], dry_run: bool = True) -> bool:
    """
    Import matched species as sightings into their corresponding Birders surveys.

    Args:
        note_data_list: List of NoteData with matched species
        dry_run: If True, preview only; if False, write to database

    Returns:
        True if successful, False otherwise
    """
    if not note_data_list:
        logger.info("No data to import.")
        return True

    logger.info("\n" + "="*80)
    logger.info("IMPORT SIGHTINGS")
    logger.info("="*80)

    with get_db_cursor() as cursor:
        # Preview: find surveys, check for existing sightings, and build import plan
        import_plan = []  # List of (survey_id, date, new_matches, skipped_matches)
        skipped_dates = []

        for note_data in sorted(note_data_list, key=lambda x: x.date):
            survey_id = find_birders_survey(cursor, note_data.date)
            if survey_id is None:
                skipped_dates.append(note_data.date)
                logger.warning(f"  No {SURVEY_TYPE_NAME} survey found for {note_data.date} - skipping")
                continue

            # Check which species already have sightings in this survey
            cursor.execute(
                "SELECT species_id FROM sighting WHERE survey_id = %s",
                (survey_id,)
            )
            existing_species_ids = {row[0] for row in cursor.fetchall()}

            new_matches = []
            skipped_matches = []
            for match in note_data.species_list:
                if match.db_id in existing_species_ids:
                    skipped_matches.append(match)
                else:
                    new_matches.append(match)

            import_plan.append((survey_id, note_data.date, new_matches, skipped_matches))

        # Log the plan
        total_new = sum(len(new) for _, _, new, _ in import_plan)
        total_skipped = sum(len(skipped) for _, _, _, skipped in import_plan)
        logger.info(f"\nSurveys matched: {len(import_plan)}")
        logger.info(f"Sightings to create: {total_new}")
        logger.info(f"Sightings already present (skipped): {total_skipped}")
        if skipped_dates:
            logger.info(f"Dates skipped (no survey): {len(skipped_dates)}")
            for d in skipped_dates:
                logger.info(f"  - {d}")

        logger.info("\nImport plan:")
        for survey_id, survey_date, new_matches, skipped_matches in import_plan:
            if not new_matches and not skipped_matches:
                continue
            logger.info(f"  Survey {survey_id} ({survey_date}):")
            for match in new_matches:
                logger.info(f"    + {match.db_name} x{match.count} (species_id={match.db_id})")
            for match in skipped_matches:
                logger.info(f"    ~ {match.db_name} x{match.count} (species_id={match.db_id}) ALREADY EXISTS")

        if dry_run:
            logger.info("\n" + "="*80)
            logger.info("DRY RUN - no changes made. Run with --no-dry-run to apply.")
            logger.info("="*80)
            return True

        # Apply: create sightings
        logger.info("\nApplying to database...")
        sightings_created = 0

        for survey_id, survey_date, new_matches, _ in import_plan:
            for match in new_matches:
                cursor.execute("""
                    INSERT INTO sighting (survey_id, species_id, count)
                    VALUES (%s, %s, %s)
                """, (survey_id, match.db_id, match.count))
                sightings_created += 1

        logger.info(f"\nCreated {sightings_created} sightings across {len(import_plan)} surveys")
        logger.info("="*80)

    return True


def main(dry_run: bool = True, verbose: bool = False):
    """Main script execution."""
    mode = "DRY RUN" if dry_run else "LIVE"
    logger.info("="*80)
    logger.info(f"EXTRACT MAMMALS/REPTILES FROM BIRD SURVEY NOTES - {mode}")
    logger.info("="*80)

    try:
        note_data_list, unmatched_list = read_excel_and_extract_notes(verbose=verbose)
        log_results(note_data_list, unmatched_list)
        import_sightings(note_data_list, dry_run=dry_run)
        return 0
    except Exception as e:
        logger.error(f"Error: {e}", exc_info=True)
        return 1


if __name__ == "__main__":
    parser = get_arg_parser(description=__doc__)
    parser.add_argument('-v', '--verbose', action='store_true', help='Enable verbose output')
    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    sys.exit(main(dry_run=args.dry_run, verbose=args.verbose))
