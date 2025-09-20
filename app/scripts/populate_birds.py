#!/usr/bin/env python3
"""
Populate database with bird survey data from Excel.
Consolidates extract_bird_data.py + import_bird_data.py + update_conservation_status.py functionality.
"""

import pandas as pd
import psycopg2
import os
import re
from datetime import datetime, date
from typing import Dict, List, Optional

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Database connection settings
DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'localhost'),
    'port': int(os.getenv('DB_PORT', 5432)),
    'database': os.getenv('DB_NAME', 'heal_butterflies'),
    'user': os.getenv('DB_USER', 'postgres'),
    'password': os.getenv('DB_PASSWORD', 'password')
}

# Conservation status lists
RED_LIST_SPECIES = [
    'Cuckoo', 'Curlew', 'Fieldfare', 'Greenfinch', 'Grey Partridge', 'Herring Gull',
    'House Sparrow', 'Lapwing', 'Lesser Redpoll', 'Linnet', 'Mistle Thrush', 'House Martin',
    'Nightingale', 'Skylark', 'Spotted Flycatcher', 'Starling', 'Swift', 'Tree Pipit',
    'Whinchat', 'Yellowhammer', 'Marsh Tit'
]

AMBER_LIST_SPECIES = [
    'Black-headed Gull', 'Bullfinch', 'Mallard', 'Cattle Egret', 'Common Gull', 'Dunnock',
    'Great Black-backed Gull', 'Great White Egret', 'Green Sandpiper', 'Grey Wagtail',
    'Greylag Goose', 'Kestrel', 'Lesser Black-backed Gull', 'Meadow Pipit', 'Moorhen',
    'Quail', 'Redstart', 'Redwing', 'Reed Bunting', 'Rook', 'Sedge Warbler',
    'Short-eared Owl', 'Snipe', 'Song Thrush', 'Sparrowhawk', 'Stock Dove', 'Tawny Owl',
    'Wheatear', 'Whitethroat', 'Woodpigeon', 'Wren'
]

def connect_db():
    """Connect to the PostgreSQL database."""
    return psycopg2.connect(**DB_CONFIG)

def get_species_colors(excel_path: str) -> Dict[str, str]:
    """Extract species colors from Excel file for conservation status."""
    if not OPENPYXL_AVAILABLE:
        return {}
    
    color_mapping = {
        'FFD9EAD3': 'Green',
        'FFFFD966': 'Amber',  
        'FFF4CCCC': 'Red',
        'FFFF6B6B': 'Red',
    }
    
    species_colors = {}
    
    try:
        wb = load_workbook(excel_path, data_only=False)
        
        for sheet_name in ['Heal Somerset bird list 2025', 'Heal Somerset bird list 2024']:
            if sheet_name not in wb.sheetnames:
                continue
                
            ws = wb[sheet_name]
            
            for row in range(5, min(120, ws.max_row + 1)):
                cell = ws[f'A{row}']
                if cell.value and isinstance(cell.value, str):
                    species_name = str(cell.value).strip()
                    if species_name and species_name not in ['Species', 'Amber List', 'Green List', 'Red List']:
                        normalized_name = normalize_species_name(species_name)
                        fill = cell.fill
                        if fill and fill.patternType and fill.fgColor and hasattr(fill.fgColor, 'rgb'):
                            color_rgb = fill.fgColor.rgb
                            if color_rgb and color_rgb in color_mapping:
                                species_colors[normalized_name] = color_mapping[color_rgb]
        
        wb.close()
        
    except Exception as e:
        print(f"Warning: Could not extract colors from Excel: {e}")
    
    return species_colors

def normalize_field_section(field_section: str) -> str:
    """Normalize field section names."""
    if pd.isna(field_section) or not isinstance(field_section, str):
        return 'Southern'  # Default

    field_section = field_section.lower().strip()

    if 'south' in field_section:
        return 'Southern'
    elif 'north' in field_section:
        return 'Northern'
    elif 'east' in field_section or 'village' in field_section:
        return 'Eastern'
    else:
        return 'Southern'

def normalize_species_name(species_name: str) -> str:
    """Normalize species names to handle duplicates and variants."""
    if pd.isna(species_name) or not isinstance(species_name, str):
        return species_name

    species_name = species_name.strip()

    # Species name mappings to normalize duplicates
    species_mappings = {
        'Heron': 'Grey Heron',
        'Domestic Mallard': 'Mallard',
        'Mallard Duck': 'Mallard',
        'Mallard duck': 'Mallard',
        'Partridge,red leg': 'Red-legged Partridge'
    }

    return species_mappings.get(species_name, species_name)

def extract_surveyors_from_text(surveyor_text: str) -> List[str]:
    """Extract individual surveyor names from text."""
    if pd.isna(surveyor_text) or not isinstance(surveyor_text, str):
        return []
    
    surveyor_text = surveyor_text.strip()
    if not surveyor_text or surveyor_text.lower() in ['nan', 'surveyors:']:
        return []
    
    # Apply name mapping before processing
    name_mapping = {
        "MarkP": "Mark",
        "Sarah Macfarlane": "Sarah",
        "Tom White": "Tom", 
        "Oli Haill": "Oli",
        "Izzy De Wattripont": "Izzy",
        "Izzy de Wattripont": "Izzy",  # Handle lowercase variant
        "Jeremy Millward": "Jeremy",
        "Lucy Carter": "Lucy",
        "Sam Wilson": "Sam",
        "Tegan Newman": "Tegan",
        "Mark Sam": "Mark & Sam"
    }
    
    # Apply mapping
    for old_name, new_name in name_mapping.items():
        surveyor_text = surveyor_text.replace(old_name, new_name)
    
    surveyors = re.split(r'[,&]+|and(?=\s)', surveyor_text)
    
    cleaned_surveyors = []
    for surveyor in surveyors:
        surveyor = surveyor.strip()
        surveyor = re.sub(r'^(mark\s*pollock|dr\.?\s*)', '', surveyor, flags=re.IGNORECASE)
        if surveyor and len(surveyor) > 1:
            cleaned_surveyors.append(surveyor.title())
    
    return cleaned_surveyors

def parse_date_value(date_val) -> Optional[date]:
    """Parse various date formats from Excel."""

    if date_val == datetime(2002, 3, 15):
        return date(2025, 3, 15)

    if pd.isna(date_val):
        return None
    
    if isinstance(date_val, datetime):
        return date_val.date()
    
    if isinstance(date_val, str):
        date_val = date_val.strip()
        if not date_val or 'e.g' in date_val.lower():
            return None
        
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
            try:
                return datetime.strptime(date_val, fmt).date()
            except ValueError:
                continue
        
        date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', date_val)
        if date_match:
            try:
                return datetime.strptime(date_match.group(1), '%d/%m/%Y').date()
            except ValueError:
                try:
                    return datetime.strptime(date_match.group(1), '%m/%d/%Y').date()
                except ValueError:
                    pass
    
    return None

def setup_transects(cursor):
    """Set up bird transects (field sections)."""
    print("Setting up bird transects...")
    
    # Clear existing transects for birds and add field sections
    cursor.execute("DELETE FROM transect WHERE name IN ('Southern', 'Northern', 'Eastern')")
    
    transects = [
        (1, 'Southern'),  # Use high numbers to avoid conflicts
        (2, 'Northern'),
        (3, 'Eastern')
    ]
    
    for number, name in transects:
        # Insert new transect (should work since we deleted above)
        cursor.execute(
            "INSERT INTO transect (number, name, type) VALUES (%s, %s, 'bird')",
            (number, name)
        )
    
    print(f"✅ Set up {len(transects)} bird field sections")

def get_or_create_surveyor(cursor, name: str) -> int:
    """Get or create a surveyor and return the ID."""
    name_parts = name.strip().split()
    first_name = name_parts[0] if name_parts else name
    last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''
    
    cursor.execute(
        "SELECT id FROM surveyor WHERE first_name = %s AND last_name = %s",
        (first_name, last_name)
    )
    result = cursor.fetchone()
    
    if result:
        return result[0]
    
    cursor.execute(
        "INSERT INTO surveyor (first_name, last_name) VALUES (%s, %s) RETURNING id",
        (first_name, last_name)
    )
    return cursor.fetchone()[0]

def get_or_create_species(cursor, name: str, conservation_status: str = None) -> int:
    """Get or create a bird species and return the ID."""
    cursor.execute("SELECT id FROM species WHERE name = %s AND type = 'bird'", (name,))
    result = cursor.fetchone()
    
    if result:
        if conservation_status:
            cursor.execute(
                "UPDATE species SET conservation_status = %s WHERE id = %s",
                (conservation_status, result[0])
            )
        return result[0]
    
    cursor.execute(
        "INSERT INTO species (name, conservation_status, type) VALUES (%s, %s, 'bird') RETURNING id",
        (name, conservation_status)
    )
    return cursor.fetchone()[0]

def extract_and_import_birds(excel_file: str):
    """Extract bird data from Excel and import to database."""
    
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        print(f"   Looking for: {os.path.abspath(excel_file)}")
        print(f"   Current directory: {os.getcwd()}")
        return
    
    conn = connect_db()
    cursor = conn.cursor()
    
    try:
        # Clear existing bird data
        print("Clearing existing bird data...")
        cursor.execute("DELETE FROM sighting WHERE species_id IN (SELECT id FROM species WHERE type = 'bird')")
        cursor.execute("DELETE FROM survey_surveyor WHERE survey_id IN (SELECT id FROM survey WHERE id NOT IN (SELECT DISTINCT survey_id FROM sighting WHERE species_id IN (SELECT id FROM species WHERE type = 'butterfly')))")
        cursor.execute("DELETE FROM survey WHERE id NOT IN (SELECT DISTINCT survey_id FROM sighting)")
        cursor.execute("DELETE FROM species WHERE type = 'bird'")
        
        # Set up transects for birds
        setup_transects(cursor)
        
        # Get species colors
        print("Extracting species colors...")
        species_colors = get_species_colors(excel_file)
        
        # Process Excel sheets
        xl = pd.ExcelFile(excel_file)
        all_species = set()
        all_surveys = []
        all_sightings = []
        all_surveyors = set()
        
        for sheet_name in ['Heal Somerset bird list 2025', 'Heal Somerset bird list 2024']:
            if sheet_name not in xl.sheet_names:
                continue
            
            print(f"Processing {sheet_name}...")
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            year = 2025 if '2025' in sheet_name else 2024
            
            # Extract species from column A
            species_col = df.iloc[:, 0]
            for species_name in species_col:
                if pd.notna(species_name) and isinstance(species_name, str):
                    species_name = species_name.strip()
                    if species_name and species_name not in ['Species', 'Amber List', 'Green List', 'Red List']:
                        normalized_name = normalize_species_name(species_name)
                        all_species.add(normalized_name)
            
            # Extract survey data and sightings
            dates_row = df.iloc[1, :].values
            field_sections_row = df.iloc[2, :].values
            surveyors_row = df.iloc[0, :].values
            
            for col_idx, date_val in enumerate(dates_row):
                parsed_date = parse_date_value(date_val)
                if not parsed_date:
                    continue
                
                field_section = normalize_field_section(str(field_sections_row[col_idx]) if col_idx < len(field_sections_row) else None)
                surveyors_text = surveyors_row[col_idx] if col_idx < len(surveyors_row) else None
                surveyor_names = extract_surveyors_from_text(str(surveyors_text))
                all_surveyors.update(surveyor_names)
                
                survey_data = {
                    'date': parsed_date,
                    'field_section': field_section,
                    'surveyors': surveyor_names,
                    'col_idx': col_idx
                }
                all_surveys.append(survey_data)
                
                # Extract sightings for this column
                for row_idx in range(5, len(species_col)):
                    species_name = species_col.iloc[row_idx] if row_idx < len(species_col) else None
                    if pd.isna(species_name) or not isinstance(species_name, str):
                        continue

                    species_name = species_name.strip()
                    normalized_name = normalize_species_name(species_name)
                    if normalized_name not in all_species:
                        continue
                    
                    if col_idx < len(df.columns) and row_idx < len(df):
                        count_value = df.iloc[row_idx, col_idx]
                        if pd.notna(count_value) and count_value != 0:
                            try:
                                count = int(float(str(count_value)))
                                if count > 0:
                                    all_sightings.append({
                                        'species_name': normalized_name,
                                        'date': parsed_date,
                                        'field_section': field_section,
                                        'count': count,
                                        'surveyors': surveyor_names
                                    })
                            except (ValueError, TypeError):
                                continue
        
        print(f"Found {len(all_species)} species, {len(all_surveys)} surveys, {len(all_sightings)} sightings")
        
        # Import species with conservation status
        print("Importing species...")
        species_ids = {}
        for species_name in all_species:
            # Determine conservation status
            conservation_status = species_colors.get(species_name)
            if not conservation_status:
                if species_name in RED_LIST_SPECIES:
                    conservation_status = 'Red'
                elif species_name in AMBER_LIST_SPECIES:
                    conservation_status = 'Amber'
                else:
                    conservation_status = 'Green'
            
            species_id = get_or_create_species(cursor, species_name, conservation_status)
            species_ids[species_name] = species_id
        
        # Import surveyors
        print("Importing surveyors...")
        surveyor_ids = {}
        for surveyor_name in all_surveyors:
            if surveyor_name.lower() not in ['et al', 'nan', '']:
                surveyor_id = get_or_create_surveyor(cursor, surveyor_name)
                surveyor_ids[surveyor_name] = surveyor_id
        
        # Get transect IDs for bird transects
        cursor.execute("SELECT id, name FROM transect WHERE name IN ('Southern', 'Northern', 'Eastern') AND type = 'bird'")
        transect_ids = {name: id for id, name in cursor.fetchall()}
        
        # Import surveys and sightings
        print("Importing surveys and sightings...")
        survey_lookup = {}  # (date, field_section) -> survey_id
        
        for survey_data in all_surveys:
            # Create survey
            cursor.execute("""
                INSERT INTO survey (date, notes, start_time, end_time, sun_percentage, temperature_celsius, conditions_met, type)
                VALUES (%s, %s, NULL, NULL, NULL, NULL, FALSE, 'bird')
                RETURNING id
            """, (survey_data['date'], None))
            survey_id = cursor.fetchone()[0]
            
            # Link surveyors
            for surveyor_name in survey_data['surveyors']:
                if surveyor_name in surveyor_ids:
                    cursor.execute("""
                        INSERT INTO survey_surveyor (survey_id, surveyor_id)
                        VALUES (%s, %s) ON CONFLICT (survey_id, surveyor_id) DO NOTHING
                    """, (survey_id, surveyor_ids[surveyor_name]))
            
            survey_lookup[(survey_data['date'], survey_data['field_section'])] = survey_id
        
        # Import sightings
        sightings_imported = 0
        for sighting in all_sightings:
            survey_key = (sighting['date'], sighting['field_section'])
            if survey_key in survey_lookup:
                survey_id = survey_lookup[survey_key]
                species_id = species_ids[sighting['species_name']]
                transect_id = transect_ids[sighting['field_section']]
                
                cursor.execute("""
                    INSERT INTO sighting (survey_id, species_id, transect_id, count)
                    VALUES (%s, %s, %s, %s)
                """, (survey_id, species_id, transect_id, sighting['count']))
                sightings_imported += 1
        
        conn.commit()
        
        # Print summary
        cursor.execute("SELECT COUNT(*) FROM species WHERE type = 'bird'")
        species_count = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM survey WHERE id IN (SELECT DISTINCT survey_id FROM sighting WHERE species_id IN (SELECT id FROM species WHERE type = 'bird'))")
        survey_count = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM sighting WHERE species_id IN (SELECT id FROM species WHERE type = 'bird')")
        sighting_count = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT conservation_status, COUNT(*) 
            FROM species 
            WHERE type = 'bird' 
            GROUP BY conservation_status 
            ORDER BY conservation_status
        """)
        status_counts = cursor.fetchall()
        
        print(f"\n🐦 Bird data import completed!")
        print(f"   - {species_count} species")
        print(f"   - {survey_count} surveys") 
        print(f"   - {sighting_count} sightings")
        print("Conservation status breakdown:")
        for status, count in status_counts:
            print(f"   - {status}: {count} species")
        
    except Exception as e:
        print(f"❌ Import failed: {e}")
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.close()

if __name__ == "__main__":
    print("DEBUG: Starting populate_birds.py script")
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_file = os.path.join(script_dir, "data", "Bird surveys - Heal Somerset - volunteers.xlsx")
    print(f"DEBUG: Excel file path: {excel_file}")
    extract_and_import_birds(excel_file)